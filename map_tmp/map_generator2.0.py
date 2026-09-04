from openpyxl import load_workbook
import json

wb = load_workbook("map.xlsx")
ws = wb.active

start_loc = (int(input("x:")), int(input("y:")))
end_loc = (int(input("x:")), int(input("y:")))
offload = (start_loc[0] - 6, start_loc[1] - 2)
route_mapping_relation = {"I": "impasse", "S": "straight",
                    "T": "turn", "H": "three",
                    "F": "four", "L": "floor",
                    }
arrow_color_mapping_relation = {"B": "blue", "R": "red", "G": "green", "Y": "yellow"}
arrow_type_mapping_relation = {"S": "straight", "T": "turn"}

map_dict = {"start":[], "end":[], "impasse":[], "straight":[],"turn":[],"three":[],"four":[],"floor":[], "arrow":[]}

def x_loc(n):
    return (int(n) - offload[0]) * 100


def y_loc(n):
    return (int(n) - offload[1]) * 100 + 55


def a(n):
    return int(n) * 90

def arrow_offload(loc, n):
    if n == 0:
        return (loc[0] - 10, loc[1] - 10)
    elif n == 1:
        return (loc[0] + 10, loc[1] - 10)
    elif n == 2:
        return (loc[0] - 10, loc[1] + 10)
    elif n == 3:
        return (loc[0] + 10, loc[1] + 10)


map_dict["start"] = [x_loc(start_loc[0]), y_loc(start_loc[1])]
map_dict["end"] = [x_loc(end_loc[0]), y_loc(end_loc[1])]

#(column,row,angle,[type],[bool])

for row in ws.iter_rows():
    for cell in row:
        if cell.value is None:
            continue
        t = cell.value
        t = t.split("_")
        type = route_mapping_relation[t[0][0]]
        x = x_loc(cell.column)
        y = y_loc(cell.row)
        ang = a(t[0][1])
        map_dict[type].append([x, y, ang])
        if len(t) == 2:
            ars = t[1].split(",")
            for ar in ars:
                type = arrow_color_mapping_relation[ar[0]] + "_" + arrow_type_mapping_relation[ar[1]]
                ang = a(ar[2])
                ax, ay = arrow_offload((x,y),int(ar[3]))
                if len(ar) == 5:
                    flip = bool(int(ar[4]))
                    map_dict["arrow"].append([ax, ay, ang, type, flip])
                else:
                    map_dict["arrow"].append([ax, ay, ang, type])
print(map_dict)
with open("map.json", "w") as f:
    json.dump(map_dict, f)
print("written")